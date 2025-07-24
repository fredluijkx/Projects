# ========================================
# 🎟️ FTO AH OVERIGE WINKELSCRAPER - HEADLESS
# Versie: 1.3b – Fred Luijkx, Netflex BV
# ========================================

import time
import pandas as pd
from datetime import datetime
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font
import smtplib
from email.message import EmailMessage
import os
from dotenv import load_dotenv
import sys

# ===============================
# 📍 Lokale paden naar Chrome en Chromedriver
chrome_path = r"C:\chrome-testing\chrome-win64\chrome.exe"
chromedriver_path = r"C:\chrome-testing\chromedriver-win64\chromedriver.exe"

# -------------------
# Helper voor PyInstaller (mag blijven)
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# -------------------
# .env inladen
dotenv_path = resource_path(".env")
load_dotenv(dotenv_path="C:/scrapers/.env")


EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")  # Komma-separated

if not EMAIL_USER or not EMAIL_PASS or not EMAIL_RECEIVER:
    print("[FATAL] Vul het bestand .env aan met: EMAIL_USER, EMAIL_PASSWORD, EMAIL_RECEIVER. Script stopt nu.")
    sys.exit(1)

EMAIL_RECEIVERS = [r.strip() for r in EMAIL_RECEIVER.split(",")]

# -------------------
# 📧 Excel mailen
def verzend_excel_via_mail(bestandspad, ontvangers, afzender, wachtwoord):
    msg = EmailMessage()
    msg["Subject"] = "🎟️ FTO – AH scraperresultaten"
    msg["From"] = afzender
    msg["To"] = ", ".join(ontvangers)
    msg.set_content("Zie bijlage voor de gescrapete producten uit Albert Heijn.")

    with open(bestandspad, "rb") as f:
        file_data = f.read()
        file_name = os.path.basename(bestandspad)

    msg.add_attachment(file_data, maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=file_name)

    with smtplib.SMTP("smtp.office365.com", 587) as smtp:
        smtp.starttls()
        smtp.login(afzender, wachtwoord)
        smtp.send_message(msg)

# -------------------
# 📅 Bestandsnaam
filename = f"ah_overige_{datetime.today().strftime('%Y-%m-%d')}.xlsx"

# -------------------
# 🔁 Scrape-functies

# Zet hier je volledige product_links en klant_namen lijsten!
product_links = [
    "https://www.ah.nl/producten/product/wi449947/fairtrade-original-kokosmelk",
    "https://www.ah.nl/producten/product/wi577746/fairtrade-original-kokosmelk-light",
    "https://www.ah.nl/producten/product/wi368572/fairtrade-original-kokosmelk-biologisch",
    "https://www.ah.nl/producten/product/wi164936/fairtrade-original-kokosmelk",
    "https://www.ah.nl/producten/product/wi516868/fairtrade-original-organic-coconut-milk",
    "https://www.ah.nl/producten/product/wi161932/ah-kokosmelk",
    "https://www.ah.nl/producten/product/wi449942/ah-biologisch-kokosmelk",
    "https://www.ah.nl/producten/product/wi553458/ah-biologisch-kokosmelk",
    "https://www.ah.nl/producten/product/wi553395/ah-kokosmelk",
    "https://www.ah.nl/producten/product/wi553394/ah-kokosmelk-light",
    "https://www.ah.nl/producten/product/wi212984/go-tan-coconut-milk",
    "https://www.ah.nl/producten/product/wi317636/go-tan-kokosmelk",
    "https://www.ah.nl/producten/product/wi388172/go-tan-bio-kokosmelk",
    "https://www.ah.nl/producten/product/wi482347/go-tan-easy-coconut-milk-lower-in-fat",   
    "https://www.ah.nl/producten/product/wi199251/fairtrade-original-kruidenpasta-groene-curry",
    "https://www.ah.nl/producten/product/wi199252/fairtrade-original-kruidenpasta-rode-curry",
    "https://www.ah.nl/producten/product/wi368566/fairtrade-original-kruidenpasta-gele-curry",
    "https://www.ah.nl/producten/product/wi581525/ah-boemboe-rode-curry",
    "https://www.ah.nl/producten/product/wi188932/koh-thai-rode-currypasta",
    "https://www.ah.nl/producten/product/wi209796/koh-thai-gele-currypasta",
    "https://www.ah.nl/producten/product/wi450167/koh-thai-groene-currypasta",
    "https://www.ah.nl/producten/product/wi553451/conimex-thaise-groene-curry",
    "https://www.ah.nl/producten/product/wi461455/fairtrade-original-hagelslag-puur",
    "https://www.ah.nl/producten/product/wi493645/de-ruijter-chocoladehagel-puur",
    "https://www.ah.nl/producten/product/wi212294/venz-hagelslag-puur",
    "https://www.ah.nl/producten/product/wi1608/ah-hagelslag-puur",
    "https://www.ah.nl/producten/product/wi563631/cote-d-or-chocolade-hagelslag-puur",
    "https://www.ah.nl/producten/product/wi168746/cereal-hagelslag-puur-minder-suikers",
    "https://www.ah.nl/producten/product/wi127402/blooker-cacaopoeder",
    "https://www.ah.nl/producten/product/wi225021/raw-organic-food-cacao-poeder",
    "https://www.ah.nl/producten/product/wi591158/fairtrade-original-cacaopoeder",
    "https://www.ah.nl/producten/product/wi388168/fairtrade-original-thaise-rijst-noedels-witte-rijst",
    "https://www.ah.nl/producten/product/wi388268/fairtrade-original-thaise-rijst-noedels-zilvervliesrijst",
    "https://www.ah.nl/producten/product/wi553339/fairtrade-original-organic-mihoen-rice-vermicelli",
    "https://www.ah.nl/producten/product/wi376579/conimex-rijstnoedels-5mm",
    "https://www.ah.nl/producten/product/wi376580/conimex-rijstnoedels-2mm",
    "https://www.ah.nl/producten/product/wi55943/go-tan-mihoen-rice-noodles",
    "https://www.ah.nl/producten/product/wi199152/go-tan-mungo-vermicelli-soe-oen",
    "https://www.ah.nl/producten/product/wi216479/koh-thai-rijst-noedels",
    "https://www.ah.nl/producten/product/wi185782/fairtrade-original-ananas-schijven-op-sap",
    "https://www.ah.nl/producten/product/wi374356/fairtrade-original-ananas-stukjes-op-eigen-sap",
    "https://www.ah.nl/producten/product/wi428748/del-monte-ananasschijven-op-sap",
    "https://www.ah.nl/producten/product/wi185582/del-monte-ananasblokjes-op-sap",
    "https://www.ah.nl/producten/product/wi857/ah-ananasschijven-op-sap",
    "https://www.ah.nl/producten/product/wi56754/ah-ananasstukjes-op-sap",
]

klant_namen = [
    "FAIRTRADE ORIGINAL 200ML",
    "FAIRTRADE ORIGINAL 200ML LIGHT",
    "FAIRTRADE ORIGINAL 270ML BIO",
    "FAIRTRADE ORIGINAL 400ML",
    "FAIRTRADE ORIGINAL 1L",
    "EIGEN MERK 200ML",
    "EIGEN MERK 200ML BIO FAIRTRADE LIGHT",
    "EIGEN MERK 400ML",
    "EIGEN MERK 500ML BIO & FAIRTRADE",
    "EIGEN MERK 500ML FAIRTRADE 17% vet",
    "EIGEN MERK 500ML FAIRTRADE light",
    "GO TAN 250ML (TETRA) 18% vet",
    "GO TAN 250ML (TETRA) BIO 18% vet",
    "GO TAN 500ML (TETRA) 8% vet",
    "FAIRTRADE ORIGINAL GROENE CURRY",
    "FAIRTRADE ORIGINAL RODE CURRY",
    "FAIRTRADE ORIGINAL GELE CURRY",
    "THAISE CURRY BAKJE",
    "KOH THAI POT  GROEN EN ROOD",
    "KOH THAI POT GEEL",
    "KOH THAI PAKJE",
    "CONIMEX ZAKJE (ROOD-GEEL-GROEN)",
    "FAIRTRADE ORIGINAL HAGELSLAG PUUR, 380G",
    "DE RUIJTER PUUR, 390G",
    "VENZ, 400G",
    "AH EIGEN MERK, 400G",
    "COTE D'OR PUUR, 200G",
    "CEREAL, 200G",
    "BLOOKER CACAOPOEDER, 250G",
    "RAW CACAOPOEDER, BIOLOGISCH, 100G",
    "FAIRTRADE ORIGINAL CACAO, 125G",
    "FAIRTRADE ORIGINAL WITTE RIJSTNOEDELS",
    "FAIRTRADE ORIGINAL ZILVERVLIES RIJSTNOEDELS",
    "FAIRTRADE ORIGINAL RIJST VERMICELLI",
    "CONIMEX MIHOEN RIJSTNOEDELS 5MM 225G",
    "CONIMEX MIHOEN RIJSTNOEDELS 2MM 225G",
    "GO TAN MIHOEN 250G",
    "GO TAN VERMICELLI GLASNOEDELS 2X50G",
    "KOH THAI RIJSTNOEDELS 220G",
    "FAIRTRADE ORIGINAL ANANASSCHIJVEN OP SAP, 565G",
    "FAIRTRADE ORIGINAL ANANASSTUKJES OP SAP, 227G",
    "DEL MONTE ANANASSCHIJVEN OP SAP, 565G",
    "DEL MONTE ANANASSTUKJES OP SAP, 227G",
    "E.M. ANANASSCHIJVEN OP SAP, 220G",
    "E.M. ANANASSTUKJES OP SAP, 227G",
]

def accept_privacy_preferences(driver):
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(text(), "Accepteren")]'))
        )
        btn.click()
        time.sleep(1)
    except Exception:
        pass

def scrape_ah_product(url):
    options = uc.ChromeOptions()
    options.binary_location = chrome_path
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--incognito")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    driver = uc.Chrome(
        options=options,
        driver_executable_path=chromedriver_path
    )
    wait = WebDriverWait(driver, 20)

    driver.get(url)
    accept_privacy_preferences(driver)

    try:
        wait.until(lambda d: d.find_element(By.CSS_SELECTOR, 'span[class*="line-clamp_root"]').text.strip() != "")
        title = driver.find_element(By.CSS_SELECTOR, 'span[class*="line-clamp_root"]').text.strip()
    except Exception:
        title = "Onbekend"

    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class*="product-card-header_unitInfo"]')))
        inhoud_text = driver.find_element(By.CSS_SELECTOR, 'div[class*="product-card-header_unitInfo"]').get_attribute("innerText")
        inhoud = inhoud_text.split("Prijs per")[0].strip()
    except Exception:
        inhoud = "Onbekend"

    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span[aria-hidden="true"]')))
        spans = driver.find_elements(By.CSS_SELECTOR, 'span[aria-hidden="true"]')
        euro = cents = None
        for span in spans:
            if span.text.strip().isdigit():
                if euro is None:
                    euro = span.text.strip()
                else:
                    cents = span.text.strip()
                    break
        prijs = f"€{euro},{cents}" if euro and cents else "Onbekend"
    except Exception:
        prijs = "Onbekend"

    driver.quit()
    return {"Productnaam": title, "Inhoud": inhoud, "Prijs": prijs, "Link": url}

# -------------------
# 🔄 Scraping loop
data = []
for link in product_links:
    print(f"[INFO] Scraping {link}")
    data.append(scrape_ah_product(link))
    time.sleep(20)

for i in range(len(data)):
    if i < len(klant_namen):
        data[i]["Productnaam"] = klant_namen[i]

print(f"[INFO] Schrijf resultaten naar {filename}")
pd.DataFrame(data).to_excel(filename, index=False)
wb = load_workbook(filename)
ws = wb.active
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.hyperlink = cell.value
        cell.font = Font(color="0000EE", underline="single")
wb.save(filename)

# -------------------
# 📧 Excel mailen, console only + automatische shutdown
try:
    verzend_excel_via_mail(
        bestandspad=filename,
        ontvangers=EMAIL_RECEIVERS,
        afzender=EMAIL_USER,
        wachtwoord=EMAIL_PASS
    )
except Exception as e:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("E-mail verzenden mislukt", f"Er ging iets mis met verzenden:\n\n{e}")
    root.destroy()
else:
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Klaar!", "De scraper is klaar en de resultaten zijn gemaild.")
    root.destroy()
    
    import subprocess
subprocess.run(["shutdown", "/s", "/t", "60"])


logging.info("Script klaar, exit nu expliciet.")
sys.exit(0)
