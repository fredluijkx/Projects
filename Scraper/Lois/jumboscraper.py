# ========================================
# 🎟️ FTO WINKELSCRAPER – JUMBO (.env, no-cert-fix)
# ========================================

import time
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import smtplib
from email.message import EmailMessage
import os
import sys
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv

# === PADEN AANPASSEN NAAR JOUW VM! ===
chrome_path = r"C:\chrome-testing\chrome-win64\chrome.exe"
chromedriver_path = r"C:\chrome-testing\chromedriver-win64\chromedriver.exe"
# ======================================

# === .env inladen ===
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

dotenv_path = resource_path(".env")
load_dotenv(dotenv_path="C:/scrapers/.env")


EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")

if not EMAIL_USER or not EMAIL_PASS or not EMAIL_RECEIVER:
    print("[FATAL] Vul het bestand .env aan met: EMAIL_USER, EMAIL_PASSWORD, EMAIL_RECEIVER. Script stopt nu.")
    sys.exit(1)
EMAIL_RECEIVERS = [r.strip() for r in EMAIL_RECEIVER.split(",")]

# === Mailfunctie (Excel) ===
def verzend_excel_via_mail(bestandspad, ontvangers, afzender, wachtwoord):
    msg = EmailMessage()
    msg["Subject"] = "🎟️ FTO – Jumbo scraperresultaten"
    msg["From"] = afzender
    msg["To"] = ", ".join(ontvangers)
    msg.set_content("Zie bijlage voor de gescrapete producten van Jumbo.")
    with open(bestandspad, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(bestandspad)
        )
    with smtplib.SMTP("smtp.office365.com", 587) as smtp:
        smtp.starttls()
        smtp.login(afzender, wachtwoord)
        smtp.send_message(msg)

# === Productlinks en klantnamen ===
product_links = [
    "https://www.jumbo.com/producten/fairtrade-original-aroma-snelfiltermaling-250-g-453622STK",
    "https://www.jumbo.com/producten/fairtrade-original-aroma-snelfiltermaling-500-g-453629STK",
    "https://www.jumbo.com/producten/fairtrade-original-mild-biologische-snelfiltermaling-250-g-453623STK",
    "https://www.jumbo.com/producten/fairtrade-orginal-decaf-snelfiltermaling-250-g-453624STK",
    "https://www.jumbo.com/producten/fairtrade-original-dark-roast-snelfiltermaling-250-g-453627STK",
    "https://www.jumbo.com/producten/fairtrade-original-goud-snelfiltermaling-250-g-453626STK",
    "https://www.jumbo.com/producten/fairtrade-original-espresso-bonen-500g-zak-623761ZK",
    "https://www.jumbo.com/producten/fairtrade-original-dark-roast-bonen-500g-624741ZK",
    "https://www.jumbo.com/producten/fairtrade-original-intense-roast-bonen-500g-624747ZK",
    "https://www.jumbo.com/producten/douwe-egberts-aroma-rood-filterkoffie-250-g-220698PAK",
    "https://www.jumbo.com/producten/douwe-egberts-aroma-rood-filterkoffie-500-g-220752PAK",
    "https://www.jumbo.com/producten/douwe-egberts-aroma-rood-grove-maling-filterkoffie-250-g-193115PAK",
    "https://www.jumbo.com/producten/douwe-egberts-aroma-rood-grove-maling-filterkoffie-500-g-112609PAK",
    "https://www.jumbo.com/producten/douwe-egberts-decafe-filterkoffie-250-g-145384DS",
    "https://www.jumbo.com/producten/douwe-egberts-decafe-filterkoffie-500-g-157736PAK",
    "https://www.jumbo.com/producten/douwe-egberts-excellent-select-filtermaling-250-g-583708STK",
    "https://www.jumbo.com/producten/jumbo-s-beste-koffie-snelfiltermaling-aroma-250-g-637243STK",
    "https://www.jumbo.com/producten/jumbo-s-beste-koffie-snelfiltermaling-aroma-500-g-637315PAK",
    "https://www.jumbo.com/producten/jumbo-s-beste-koffie-snelfiltermaling-goud-500-g-637313STK",
    "https://www.jumbo.com/producten/jumbo-snelfiltermaling-biologisch-250-g-318545ZK",
    "https://www.jumbo.com/producten/kanis-gunnink-regular-filterkoffie-500-g-865249PAK",
    "https://www.jumbo.com/producten/van-nelle-filterkoffie-2-x-250-g-331384PAK",
    "https://www.jumbo.com/producten/lavazza-qualita-rossa-gemalen-filterkoffie-250-g-224264STK",
    "https://www.jumbo.com/producten/lavazza-espresso-italiano-classico-gemalen-filterkoffie-250-g-267848PAK",
    "https://www.jumbo.com/producten/lavazza-qualita-oro-gemalen-filterkoffie-250-g-224263STK",
    "https://www.jumbo.com/producten/cafe-intencion-aromatico-250-g-filterkoffie-70307PAK",
    "https://www.jumbo.com/producten/cafe-intencion-fuerte-filterkoffie-250-g-549589ZK",
    "https://www.jumbo.com/producten/douwe-egberts-aroma-rood-bonen-500-g-51016PAK",
    "https://www.jumbo.com/producten/douwe-egberts-aroma-rood-koffiebonen-voordeelpak-1000-g-390432ZK",
    "https://www.jumbo.com/producten/douwe-egberts-espresso-bonen-500-g-140445ZK",
    "https://www.jumbo.com/producten/douwe-egberts-espresso-koffiebonen-voordeelpak-1000-g-390430ZK",
    "https://www.jumbo.com/producten/douwe-egberts-excellent-100-arabica-gold-500-g-10934DS",
    "https://www.jumbo.com/producten/douwe-egberts-intens-koffiebonen-500-g-207282STK",
    "https://www.jumbo.com/producten/douwe-egberts-excellent-mocca-500-g-68547ZK",
    "https://www.jumbo.com/producten/l-or-crema-classique-koffiebonen-500-g-187034ZK",
    "https://www.jumbo.com/producten/l-or-espresso-forza-koffiebonen-500-g-108210ZK",
    "https://www.jumbo.com/producten/l-or-espresso-fortissimo-koffiebonen-500-g-108211ZK",
    "https://www.jumbo.com/producten/l-or-espresso-onyx-koffiebonen-500-g-140129ZK",
    "https://www.jumbo.com/producten/jumbo-s-beste-koffiebonen-aroma-1-kg-637316ZK",
    "https://www.jumbo.com/producten/jumbo-s-beste-koffiebonen-espresso-intenso-1-kg-637272ZK",
    "https://www.jumbo.com/producten/jumbo-s-beste-koffiebonen-dark-roast-1-kg-637271ZK",
    "https://www.jumbo.com/producten/la-place-koffiebonen-originale-1kg-323798ZK",
    "https://www.jumbo.com/producten/la-place-koffiebonen-espresso-1-kg-323799ZK",
    "https://www.jumbo.com/producten/kanis-gunnink-medium-koffiebonen-1-kg-390431ZK",
    "https://www.jumbo.com/producten/illy-classic-roast-250-g-62456BLK",
    "https://www.jumbo.com/producten/illy-intenso-bold-roast-coffee-beans-250-g-225546STK",
    "https://www.jumbo.com/producten/lavazza-qualita-rossa-koffiebonen-1kg-576366ZK",
    "https://www.jumbo.com/producten/lavazza-qualita-oro-koffiebonen-500-g-66421PAK",
    "https://www.jumbo.com/producten/cafe-intencion-crema-aromatico-500-g-bonen-130141PAK",
    "https://www.jumbo.com/producten/cafe-intencion-espresso-intensivo-500-g-bonen-372105PAK",
    "https://www.jumbo.com/producten/fairtrade-original-kokos-melk-200-ml-562499BLK",
    "https://www.jumbo.com/producten/fairtrade-original-kokosmelk-light-200ml-627844PAK",
    "https://www.jumbo.com/producten/fairtrade-original-organic-coconut-milk-270-ml-161422BLK",
    "https://www.jumbo.com/producten/fairtrade-original-kokosmelk-400-ml-434972BLK",
    "https://www.jumbo.com/producten/fairtrade-original-biologische-kokosmelk-1-l-373594PAK",
    "https://www.jumbo.com/producten/jumbo-kokosmelk-romig-200-ml-224389PAK",
    "https://www.jumbo.com/producten/jumbo-kokosmelk-400-ml-408583BLK",
    "https://www.jumbo.com/producten/jumbo-kokosmelk-6-vet-400-ml-408584BLK",
    "https://www.jumbo.com/producten/jumbo-biologisch-light-kokosmelk-400ml-224785BLK",
    "https://www.jumbo.com/producten/jumbo-biologisch-romige-kokosmelk-400ml-224783BLK",
    "https://www.jumbo.com/producten/go-tan-kokosmelk-250ml-161809PAK",
    "https://www.jumbo.com/producten/go-tan-biologische-kokosmelk-250ml-211894STK",
    "https://www.jumbo.com/producten/go-tan-kokosmelk-8-500ml-435953PAK",
    "https://www.jumbo.com/producten/fairtrade-original-thaise-groene-curry-70-g-145735PAK",
    "https://www.jumbo.com/producten/fairtrade-original-thaise-rode-curry-70-g-145734PAK",
    "https://www.jumbo.com/producten/fairtrade-original-thaise-gele-curry-70-g-161421DS",
    "https://www.jumbo.com/producten/jumbo-boemboe-groene-curry-95-g-194781STK",
    "https://www.jumbo.com/producten/koh-thai-groene-curry-pasta-69743POT",
    "https://www.jumbo.com/producten/koh-thai-rode-curry-pasta-69747POT",
    "https://www.jumbo.com/producten/koh-thai-gele-curry-pasta-69749POT",
    "https://www.jumbo.com/producten/koh-thai-rode-curry-70-g-303217PAK",
    "https://www.jumbo.com/producten/conimex-pasta-thaise-rode-curry-90-g-575126ZK",
    "https://www.jumbo.com/producten/fairtrade-original-chocolade-hagelslag-puur-380g-628643DS",
    "https://www.jumbo.com/producten/de-ruijter-hagelslag-puur-390-g-417701DS",
    "https://www.jumbo.com/producten/venz-puur-chocolade-hagelslag-400-g-160610DS",
    "https://www.jumbo.com/producten/jumbo-chocolade-hagelslag-puur-380-g-625076DS",
    "https://www.jumbo.com/producten/cereal-minder-suikers-hagelslag-puur-200-g-565020DS"
]
klant_namen = [
    "LIRP Koffie Aroma snf, MH, 250g",
    "LIRP Koffie Aroma snf, MH, 500g",
    "LIRP Koffie Mild snf, bio, MH, 250g ",
    "LIRP Koffie Décaf snf, MH, 250g",
    "LIRP Koffie Dark Roast snf MH, 250g",
    "LIRP Koffie Goud snf, MH, 250g",
    "LIRP Koffiebonen Espresso, 500g",
    "LIRP Koffiebonen Espresso Dark Roast, 500g",
    "LIRP Koffiebonen Espr. XDR bio /vanaf wk27 '24: Intense Roast",
    "D.E. AROMA  ROOD 250G",
    "D.E. AROMA ROOD 500G",
    "D.E. AROMA ROOD GROVE MALING 250G",
    "D.E. AROMA ROOD GROVE MALING 500G",
    "D.E. DECAFE 250G",
    "D.E. DECAFE 500G",
    "D.E. EXCELLENT 5 250G",
    "JUMBO AROMA ROOD 250G",
    "JUMBO AROMA ROOD 500G",
    "JUMBO GOUD 500G",
    "JUMBO BIO 250G (softpack)",
    "KAN.&GUNNINK 500G",
    "V.NELLE SUPRA  KOFFIEKRACHT 5 2STx250g",
    "LAVAZZA QUALITA ROSSA 250G",
    "LAVAZZA ITALIANO CLASSICO ESPRESSO 250G",
    "LAVAZZA QUALITA ORO PERFECT SYMPH. ESPRESSO 250G",
    "CAFE INTENCION ECOLOGICO 250G",
    "CAFE INTENCION FUERTE 250G",
    "D.E. AROMA ROOD 500G",
    "D.E. AROMA ROOD 1.000G",
    "D.E. ESPRESSO 9 500G",
    "D.E. ESPRESSO 9 1.000G",
    "D.E. EXCELLENT 5 500G",
    "D.E. INTENS BONEN 7 500G",
    "D.E. MOCCA 7 500G",
    "L'OR CREMA CLASSIQUE 500G",
    "L'OR FORZA 500G",
    "L'OR FORTISSIMO 500G",
    "L'OR ONYX 500G",
    "JUMBO AROMA BONEN 1.000G",
    "JUMBO ESPRESSO 1.000G",
    "JUMBO DARK ROAST ESPRESSO 1.000G",
    "LA PLACE KOFFIEBONEN ORIGINALE 1.000g",
    "LA PLACE ESPRESSO 1.000g",
    "KAN.&GUNNINK MEDIUM ROAST 1.000G",
    "ILLY ESPRESSO CLASSICO 250G (BLIK)",
    "ILLY ESPRESSO DARK ROAST BONEN 250G (BLIK)",
    "LAVAZZA QUALITA ROSSA 500G",
    "LAVAZZA QUALITA ORO PERFECT SYMPH. ESPRESSO 500G",
    "CAFE INTENCION KOFFIEBONEN 500G",
    "CAFE INTENCION ESPR.BONEN 500G",
    "FAIRTRADE ORIGINAL 200ML",
    "FAIRTRADE ORIGINAL 200ML LIGHT",
    "FAIRTRADE ORIGINAL 270ML BIO",
    "FAIRTRADE ORIGINAL 400ML",
    "FAIRTRADE ORIGINAL 1L",
    "EIGEN MERK 200ML",
    "EIGEN MERK 400ML 18% (tot wk20 6% vet gemeten)",
    "EIGEN MERK 400ML 6,6% / 6,0%",
    "EIGEN MERK 400ML BIO FAIRTRADE 6%",
    "EIGEN MERK 400ML BIO FAIRTRADE ROMIG",
    "GO TAN 250ML (TETRA)",
    "GO TAN 250ML (TETRA) BIO",
    "GO TAN 500ML (TETRA)",
    "FAIRTRADE ORIGINAL GROENE CURRY",
    "FAIRTRADE ORIGINAL RODE CURRY",
    "FAIRTRADE ORIGINAL GELE CURRY",
    "EM THAISE CURRY PAKJE",
    "KOH THAI POT GROENE CURRY",
    "KOH THAI POT  RODE CURRY",
    "KOH THAI POT GELE CURRY",
    "KOH THAI PAKJE",
    "CONIMEX ZAKJE (ROOD-GEEL-GROEN)",
    "FAIRTRADE ORIGINAL HAGELSLAG PUUR, 380G",
    "DE RUIJTER PUUR, 390G",
    "VENZ, 400G",
    "JUMBO EIGEN MERK, 380G",
    "CEREAL, 200G",
    "FAIRTRADE ORIGINAL CACAOPOEDER, 125G",
]

# === Excel-bestandsnaam ===
filename = f"jumbo_producten_{datetime.today().strftime('%Y-%m-%d')}.xlsx"

# === Selenium-setup (local driver!) ===
options = uc.ChromeOptions()
options.binary_location = chrome_path
options.add_argument("--headless=new")
options.add_argument("--disable-gpu")
options.add_argument("--window-size=1920,1080")
driver = uc.Chrome(options=options, driver_executable_path=chromedriver_path)
wait = WebDriverWait(driver, 20)

# === Scraperfunctie ===
def scrape_jumbo_html(url):
    driver.get(url)
    time.sleep(2)

    naam = "Onbekend"
    inhoud = "Onbekend"
    prijs = "Onbekend"

    try:
        naam = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1"))).text.strip()
    except Exception:
        print(f"[❌] Geen productnaam gevonden bij: {url}")

    try:
        tekst = driver.page_source.lower()
        match = re.search(r'(\d+\s?(?:g|kg|ml|l))', tekst)
        if match:
            inhoud = match.group(1)
    except Exception:
        print(f"[❌] Geen inhoud gevonden bij: {url}")

    try:
        prijs = "Onbekend"
        screenreaders = driver.find_elements(By.CLASS_NAME, "screenreader-only")
        for sr in screenreaders:
            tekst = sr.text.strip()
            if tekst.startswith("Prijs:"):
                match = re.search(r"€\s?([\d,.]+)", tekst)
                if match:
                    prijs = f"€{match.group(1).replace('.',',')}"
                    break
        if prijs == "Onbekend":
            raise Exception("Geen prijs gevonden in screenreader-only")
    except Exception as e:
        print(f"[❌] Geen prijs gevonden bij: {url} ({e})")

    return {
        "Productnaam": naam,
        "Inhoud": inhoud,
        "Prijs": prijs,
        "Link": url
    }

# === Producten ophalen ===
results = []
for i, link in enumerate(product_links, 1):
    print(f"📦 ({i}/{len(product_links)}) Ophalen: {link}")
    results.append(scrape_jumbo_html(link))
    time.sleep(2)

driver.quit()

# === Klantnamen toepassen ===
for i in range(min(len(results), len(klant_namen))):
    results[i]["Productnaam"] = klant_namen[i]

# === Excel opslaan ===
pd.DataFrame(results).to_excel(filename, index=False)

wb = load_workbook(filename)
ws = wb.active
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.hyperlink = cell.value
        cell.font = Font(color="0000EE", underline="single")
wb.save(filename)

# === Mail versturen (.env) ===
try:
    verzend_excel_via_mail(
        bestandspad=filename,
        ontvangers=EMAIL_RECEIVERS,
        afzender=EMAIL_USER,
        wachtwoord=EMAIL_PASS
    )
except Exception as e:
    print(f"[ERROR] E-mail verzenden mislukt: {e}")
    sys.exit(2)

print(f"✅ HTML scraping + e-mail voltooid: {filename}")
sys.exit(0)
