# ========================================
# 📦 FTO WINKELSCRAPER – PLUS (.env, local chrome, cert-fix)
# ========================================

import time
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import sys
import smtplib
from email.message import EmailMessage
import os
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv

# === PADEN AANPASSEN AAN JOUW VM ===
chrome_path = r"C:\chrome-testing\chrome-win64\chrome.exe"
chromedriver_path = r"C:\chrome-testing\chromedriver-win64\chromedriver.exe"
# ====================================

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

dotenv_path = resource_path(".env")
load_dotenv(dotenv_path="C:/scrapers/.env")


EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")

if not EMAIL_USER or not EMAIL_PASS or not EMAIL_RECEIVER:
    print("[FATAL] Vul het bestand .env aan met: EMAIL_USER, EMAIL_PASSWORD, EMAIL_RECEIVER. Script stopt nu.")
    sys.exit(1)
EMAIL_RECEIVERS = [r.strip() for r in EMAIL_RECEIVER.split(",")]

# === Mailfunctie (Excel) ===
def verzend_excel_via_mail(bestandspad, ontvangers, afzender, wachtwoord):
    msg = EmailMessage()
    msg["Subject"] = "🎟️ FTO – PLUS scraperresultaten"
    msg["From"] = afzender
    msg["To"] = ", ".join(ontvangers)
    msg.set_content("Zie bijlage voor de gescrapete producten van PLUS.")
    with open(bestandspad, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(bestandspad)
        )
    with smtplib.SMTP("smtp.office365.com", 587) as smtp:
        smtp.starttls()
        smtp.login(afzender, wachtwoord)
        smtp.send_message(msg)

# === PLUS-productlinks en klantnamen ===
product_links = [
    "https://www.plus.nl/product/fairtrade-original-snelfilterkoffie-aroma-zak-250-g-506193",
    "https://www.plus.nl/product/fairtrade-original-snelfilterkoffie-mild-zak-250-g-709462",
    "https://www.plus.nl/product/fairtrade-original-espresso-8-koffiebonen-zak-500-g-272201",
    "https://www.plus.nl/product/fairtrade-original-dark-roast-10-koffiebonen-zak-500-g-272203",
    "https://www.plus.nl/product/fairtrade-original-intens-roast-12-koffiebonen-zak-500-g-275192",
    "https://www.plus.nl/product/douwe-egberts-aroma-rood-filterkoffie-pak-250-g-159124",
    "https://www.plus.nl/product/douwe-egberts-aroma-rood-filterkoffie-pak-500-g-159159",
    "https://www.plus.nl/product/douwe-egberts-aroma-rood-grove-maling-filterkoffie-pak-250-g-159175",
    "https://www.plus.nl/product/douwe-egberts-aroma-rood-grove-maling-filterkoffie-pak-500-g-159213",
    "https://www.plus.nl/product/douwe-egberts-decafe-cafeinevrije-filterkoffie-pak-250-g-158659",
    "https://www.plus.nl/product/douwe-egberts-decafe-cafeinevrije-filterkoffie-pak-500-g-158675",
    "https://www.plus.nl/product/douwe-egberts-aroma-variatie-excellent-filterkoffie-pak-250-g-777794",
    "https://www.plus.nl/product/plus-roodmerk-filterkoffie-fairtrade-pak-250-g-729165",
    "https://www.plus.nl/product/plus-roodmerk-filterkoffie-fairtrade-pak-500-g-729157",
    "https://www.plus.nl/product/plus-goudmerk-filterkoffie-fairtrade-pak-250-g-729171",
    "https://www.plus.nl/product/kanis-gunnink-regular-filterkoffie-pak-500-g-160092",
    "https://www.plus.nl/product/lavazza-filterkoffie-qualita-rossa-stazak-250-g-284803",
    "https://www.plus.nl/product/lavazza-espresso-classico-filterkoffie-stazak-250-g-331218",
    "https://www.plus.nl/product/bio-koffie-snelfilter-dutch-roast-zak-250-g-330782",
    "https://www.plus.nl/product/douwe-egberts-aroma-rood-koffiebonen-zak-500-g-879849",
    "https://www.plus.nl/product/douwe-egberts-aroma-rood-koffiebonen-voordeelpak-zak-1000-g-188549",
    "https://www.plus.nl/product/douwe-egberts-aroma-variaties-excellent-koffiebonen-zak-500-g-879853",
    "https://www.plus.nl/product/douwe-egberts-aroma-variaties-mocca-koffiebonen-zak-500-g-746078",
    "https://www.plus.nl/product/l-or-crema-classique-koffiebonen-zak-500-g-255336",
    "https://www.plus.nl/product/l-or-espresso-forza-koffiebonen-zak-500-g-391619",
    "https://www.plus.nl/product/l-or-espresso-fortissimo-koffiebonen-zak-500-g-391617",
    "https://www.plus.nl/product/l-or-espresso-onyx-koffiebonen-zak-500-g-383956",
    "https://www.plus.nl/product/plus-koffiebonen-roodmerk-fairtrade-stazak-500-g-395752",
    "https://www.plus.nl/product/plus-koffiebonen-roodmerk-fairtrade-stazak-1000-g-395748",
    "https://www.plus.nl/product/plus-koffiebonen-espresso-dark-fairtrade-stazak-1000-g-395754",
    "https://www.plus.nl/product/kanis-gunnink-medium-koffiebonen-zak-1000-g-587980",
    "https://www.plus.nl/product/illy-koffiebonen-blik-250-g-616149",
    "https://www.plus.nl/product/illy-donker-koffiebonen-blik-250-g-761615",
    "https://www.plus.nl/product/lavazza-qualita-rossa-bonen-stazak-1000-g-550321",
    "https://www.plus.nl/product/lavazza-caffe-espresso-d-oro-koffiebonen-stazak-500-g-735108",
    "https://www.plus.nl/product/bio-espressobonen-fairtrade-stazak-450-g-768715",
    "https://www.plus.nl/product/roots-medium-roast-bio-stazak-500-g-381933",
    "https://www.plus.nl/product/roots-dark-roast-bio-stazak-500-g-381935",
    "https://www.plus.nl/product/roots-extra-dark-roast-bio-stazak-500-g-381937",
    "https://www.plus.nl/product/fairtrade-original-kokosmelk-fairtrade-blik-200-ml-441335",
    "https://www.plus.nl/product/fairtrade-original-organic-coconut-milk-light-pak-200-ml-184183",
    "https://www.plus.nl/product/fairtrade-original-kokosmelk-fairtrade-bio-blik-270-ml-548184",
    "https://www.plus.nl/product/fairtrade-original-kokosmelk-fairtrade-blik-400-ml-257727",
    "https://www.plus.nl/product/fairtrade-original-kokosmelk-bio-pak-1000-ml-969797",
    "https://www.plus.nl/product/sum-en-sam-kokosmelk-pak-200-ml-574321",
    "https://www.plus.nl/product/sum-en-sam-kokosmelk-6-vet-blik-400-ml-223967",
    "https://www.plus.nl/product/go-tan-kokosmelk-pak-250-ml-323221",
    "https://www.plus.nl/product/go-tan-kokosmelk-creamy-blik-400-ml-561879",
    "https://www.plus.nl/product/go-tan-kokosmelk-biologisch-pak-250-ml-196038",
    "https://www.plus.nl/product/go-tan-kokosmelk-biologisch-blik-400-ml-179951",
    "https://www.plus.nl/product/go-tan-kokosmelk-pak-500-ml-681055",
    "https://www.plus.nl/product/fairtrade-original-groene-curry-kruidenpasta-fairtrade-doos-70-g-323515",
    "https://www.plus.nl/product/fairtrade-original-rode-curry-kruidenpasta-fairtrade-doos-70-g-323517",
    "https://www.plus.nl/product/fairtrade-original-gele-curry-kruidenpasta-fairtrade-doos-70-g-288069",
    "https://www.plus.nl/product/plus-boemboe-rode-curry-kuipje-95-g-257884",
    "https://www.plus.nl/product/koh-thai-curry-pasta-groen-pot-225-g-746550",
    "https://www.plus.nl/product/koh-thai-curry-pasta-geel-pot-225-g-548123",
    "https://www.plus.nl/product/koh-thai-green-curry-paste-pak-70-g-226320",
    "https://www.plus.nl/product/conimex-gele-curry-paste-stazak-90-g-682290",
    "https://www.plus.nl/product/de-ruijter-chocoladehagel-puur-doos-390-g-930848",
    "https://www.plus.nl/product/venz-chocoladehagelslag-puur-doos-400-g-391309",
    "https://www.plus.nl/product/plus-chocoladehagelslag-puur-fairtrade-doos-400-g-459607",
    "https://www.plus.nl/product/cereal-hagelslag-puur-doos-200-g-757181",
    "https://www.plus.nl/product/fairtrade-original-cacoapoeder-doos-125-g-533769",
    "https://www.plus.nl/product/raw-organic-food-cacaopoeder-biologisch-pot-100-g-501225",
    "https://www.plus.nl/product/fairtrade-original-witte-rijst-noedels-fairtrade-zak-225-g-808530",
    "https://www.plus.nl/product/fairtrade-original-bruine-rijst-noedel-zak-225-g-808532",
    "https://www.plus.nl/product/conimex-rijst-noedels-5mm-stazak-225-g-377349",
    "https://www.plus.nl/product/conimex-rijstnoodles-2mm-stazak-225-g-666320",
    "https://www.plus.nl/product/go-tan-miehoen-zak-250-g-381353",
    "https://www.plus.nl/product/go-tan-vermicelli-soe-oen-zak-100-g-281723",
    " https://www.plus.nl/product/koh-thai-rijstnoedels-stazak-200-g-383658",
    "https://www.plus.nl/product/fairtrade-original-pineapple-slices-in-own-juice-vegan-blik-565-g-157037",
    "https://www.plus.nl/product/fairtrade-original-ananas-stukjes-op-sap-fairtrade-blik-227-g-741354",
    "https://www.plus.nl/product/del-monte-ananas-schijven-1-4-siroop-blik-235-g-821029",
    "https://www.plus.nl/product/del-monte-ananas-schijven-op-sap-blik-820-g-405907",
    "https://www.plus.nl/product/del-monte-ananas-stukjes-op-sap-blik-230-g-896293",
    "https://www.plus.nl/product/plus-ananas-schijven-op-sap-blik-567-g-510443",
    "https://www.plus.nl/product/plus-ananasstukjes-op-sap-blik-227-g-510449",
]
klant_namen = [
     "LIRP Koffie Aroma snf, MH, 250g",
    "LIRP Koffie Mild snf, bio, MH, 250g ",
    "LIRP Koffiebonen Espresso, 500g",
    "LIRP Koffiebonen Espresso Dark Roast, 500g",
    "LIRP Koffiebonen Espresso INTENS Roast (vanaf eind 2024)",
    "D.E. AROMA  ROOD 250G",
    "D.E. AROMA ROOD 500G",
    "D.E. AROMA ROOD GROVE MALING 250G",
    "D.E. AROMA ROOD GROVE MALING 500G",
    "D.E. DECAFE 250G",
    "D.E. DECAFE 500G",
    "D.E. EXCELLENT 5 250G",
    "PLUS AROMA ROOD 250G (FAIRTRADE)",
    "PLUS AROMA ROOD 500G (FAIRTRADE)",
    "PLUS GOUD 250G (FAIRTRADE)",
    "KAN.&GUNNINK 500G",
    "LAVAZZA QUALITA ROSSA 250G",
    "LAVAZZA ITALIANO CLASSICO ESPRESSO 250G",
    "BIO+ AROMA/DUTCH ROAST 250G",
    "D.E. AROMA ROOD 500G",
    "D.E. AROMA ROOD 1.000G",
    "D.E. EXCELLENT 5 500G",
    "D.E. MOCCA 7 500G",
    "L'OR CREMA CLASSIQUE 500G",
    "L'OR FORZA 500G",
    "L'OR FORTISSIMO 500G",
    "L'OR ONYX 500G",
    "PLUS AROMA BONEN 500G (FAIRTRADE)",
    "PLUS AROMA BONEN 1.000G (FAIRTRADE)",
    "PLUS DARK ROAST ESPRESSO 9 1.000G (FAIRTRADE)",
    "KAN.&GUNNINK MEDIUM ROAST 1.000G",
    "ILLY ESPRESSO CLASSICO 250G (BLIK)",
    "ILLY ESPRESSO DARK ROAST BONEN 250G (BLIK)",
    "LAVAZZA QUALITA ROSSA 1KG",
    "LAVAZZA QUALITA ORO PERFECT SYMPH. ESPRESSO 500G",
    "BIO+ ESPRESSO 450G",
    "ROOTS MEDIUM ROAST 500G",
    "ROOTS DARK ROAST 500G",
    "ROOTS EXTRA DARK ROAST 500G",
    "FAIRTRADE ORIGINAL 200ML",
    "FAIRTRADE ORIGINAL 200ML LIGHT",
    "FAIRTRADE ORIGINAL 270ML BIO",
    "FAIRTRADE ORIGINAL 400ML",
    "FAIRTRADE ORIGINAL 1L",
    "SUM&SAM 200ML 18% vet",
    "SUM&SAM 400ML 6% vet",
    "GO TAN 250ML (TETRA)",
    "GO TAN 400ML",
    "GO TAN 250ML (TETRA) BIO",
    "GO TAN 400ML BIO",
    "GO TAN 500ML (TETRA)",
    "FAIRTRADE ORIGINAL GROENE CURRY",
    "FAIRTRADE ORIGINAL RODE CURRY",
    "FAIRTRADE ORIGINAL GELE CURRY",
    "PLUS KUIPJE, 95GR",
    "KOH THAI POT ROOD EN GROEN",
    "KOH THAI POT  GEEL",
    "KOH THAI PAKJE",
    "CONIMEX ZAKJE (ROOD-GEEL-GROEN)",
    "DE RUIJTER PUUR, 390G",
    "VENZ, 400G",
    "PLUS EIGEN MERK, 400G",
    "CEREAL, 200G",
    "FAIRTRADE ORIGINAL CACAOPOEDER, 125G",
    "RAW CACAOPOEDER, BIOLOGISCH, 100G",
    "FAIRTRADE ORIGINAL WITTE RIJSTNOEDELS",
    "FAIRTRADE ORIGINAL ZILVERVLIES RIJSTNOEDELS",
    "CONIMEX MIHOEN RIJSTNOEDELS 5MM 225G",
    "CONIMEX MIHOEN RIJSTNOEDELS 2MM 225G",
    "GO TAN MIHOEN 250G",
    "GO TAN VERMICELLI GLASNOEDELS 2X50G",
    "KOH THAI RIJSTNOEDELS 200G",
    "FAIRTRADE ORIGINAL ANANASSCHIJVEN OP SAP, 565G",
    "FAIRTRADE ORIGINAL ANANASSTUKJES OP SAP, 227G",
    "DEL MONTE ANANASSCHIJVEN 1/4 SIROOP, 235G",
    "DEL MONTE ANANASSCHIJVEN OP SAP, 820G",
    "DEL MONTE ANANASBLOKJES OP SAP, 435G",
    "E.M. ANANASSCHIJVEN OP SAP, 567G",
    "E.M. ANANASSTUKJES OP SAP, 227G",
]

today = datetime.today().strftime("%Y-%m-%d")
filename = f"plus_scraper_{today}.xlsx"

# === Selenium driver setup (no cert issues) ===
options = uc.ChromeOptions()
options.binary_location = chrome_path
options.add_argument('--disable-gpu')
options.add_argument('--window-size=1920,1080')
options.add_argument('--headless=new')

try:
    driver = uc.Chrome(options=options, driver_executable_path=chromedriver_path)
    wait = WebDriverWait(driver, 30)
except Exception as e:
    print(f"[FATAL] Fout bij opstarten van browser: {e}")
    sys.exit(1)

def accept_cookies():
    try:
        accept_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"Accepteer")]'))
        )
        accept_btn.click()
        time.sleep(2)
    except:
        pass

def scrape_plus_product(url):
    driver.get(url)
    time.sleep(3)
    accept_cookies()
    time.sleep(3)

    title = "Onbekend"
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1')))
        title = driver.find_element(By.CSS_SELECTOR, 'h1').text.strip()
    except:
        try:
            title = driver.title.split(" |")[0].strip()
        except:
            print(f"[❌] Geen productnaam: {url}")

    inhoud = "Onbekend"
    try:
        spans = driver.find_elements(By.CSS_SELECTOR, 'span.place')
        for span in spans:
            match = re.search(r'(\d+\s?(?:g|kg|ml|l))', span.text.lower())
            if match:
                inhoud = match.group(1)
                break
    except:
        print(f"[❌] Geen inhoud gevonden: {url}")

    prijs = "Onbekend"
    try:
        euro = wait.until(EC.presence_of_element_located((By.ID, 'b4-b2-PriceInteger'))).text.strip().replace(".", "")
        cent = wait.until(EC.presence_of_element_located((By.ID, 'b4-b2-PriceDecimals'))).text.strip()
        prijs = f"€{euro},{cent}"
    except:
        try:
            prijs_container = driver.find_element(By.CLASS_NAME, 'product-header-price')
            match = re.search(r'€\d+,\d+', prijs_container.text)
            if match:
                prijs = match.group()
        except:
            print(f"[❌] Geen prijs gevonden: {url}")

    return {
        "Productnaam": title,
        "Inhoud": inhoud,
        "Prijs": prijs,
        "Link": url
    }

# === Scrapen zonder tqdm ===
results = []
for i, link in enumerate(product_links, 1):
    print(f"🛒 ({i}/{len(product_links)}) Ophalen: {link}")
    results.append(scrape_plus_product(link))
    time.sleep(15)

driver.quit()

# === Klantnamen toepassen ===
for i in range(min(len(results), len(klant_namen))):
    results[i]["Productnaam"] = klant_namen[i]

# === Excel opslaan ===
pd.DataFrame(results).to_excel(filename, index=False)

wb = load_workbook(filename)
ws = wb.active
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.hyperlink = cell.value
        cell.font = Font(color="0000EE", underline="single")
wb.save(filename)

# === Mail verzenden (.env) ===
try:
    verzend_excel_via_mail(
        bestandspad=filename,
        ontvangers=EMAIL_RECEIVERS,
        afzender=EMAIL_USER,
        wachtwoord=EMAIL_PASS
    )
except Exception as e:
    print(f"[ERROR] E-mail verzenden mislukt: {e}")
    sys.exit(2)

print(f"✅ PLUS scraping en e-mail voltooid: {filename}")
sys.exit(0)
