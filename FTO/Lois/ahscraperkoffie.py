# ========================================
# 📦 FTO KOFFIE-SCRAPER – Versie 1.5 
# Auteur: Fred Luijkx – Netflex BV
# ========================================

import time
import random
import pandas as pd
from datetime import datetime
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font
import smtplib
from email.message import EmailMessage
import os
import sys
from dotenv import load_dotenv

# =============== Chrome/driver paths aanpassen aan jouw VM! ===============
chrome_path = r"C:\chrome-testing\chrome-win64\chrome.exe"
chromedriver_path = r"C:\chrome-testing\chromedriver-win64\chromedriver.exe"
# ==========================================================================

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

dotenv_path = resource_path(".env")
load_dotenv(dotenv_path="C:/scrapers/.env")


EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")

if not EMAIL_USER or not EMAIL_PASS or not EMAIL_RECEIVER:
    print("[FATAL] Vul het bestand .env aan met: EMAIL_USER, EMAIL_PASSWORD, EMAIL_RECEIVER. Script stopt nu.")
    sys.exit(1)

EMAIL_RECEIVERS = [r.strip() for r in EMAIL_RECEIVER.split(",")]

def verzend_excel_via_mail(bestandspad, ontvangers, afzender, wachtwoord):
    msg = EmailMessage()
    msg["Subject"] = "📦 FTO – AH scraperresultaten (Koffie)"
    msg["From"] = afzender
    msg["To"] = ", ".join(ontvangers)
    msg.set_content("Zie bijlage voor de gescrapete koffieproducten uit Albert Heijn.")

    with open(bestandspad, "rb") as f:
        file_data = f.read()
        file_name = os.path.basename(bestandspad)

    msg.add_attachment(file_data, maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=file_name)

    with smtplib.SMTP("smtp.office365.com", 587) as smtp:
        smtp.starttls()
        smtp.login(afzender, wachtwoord)
        smtp.send_message(msg)

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
]

def accept_privacy_preferences(driver):
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(text(), "Accepteren")]'))
        )
        btn.click()
        time.sleep(1)
    except Exception:
        pass

def simulate_user_scroll(driver):
    try:
        for _ in range(random.randint(2, 4)):
            driver.execute_script("window.scrollBy(0, 300);")
            time.sleep(random.uniform(0.5, 1.5))
    except Exception:
        pass

def get_driver():
    options = uc.ChromeOptions()
    options.binary_location = chrome_path
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(f"user-agent={random.choice(USER_AGENTS)}")
    return uc.Chrome(options=options, driver_executable_path=chromedriver_path)

def scrape_ah_product(driver, url, index):
    wait = WebDriverWait(driver, 20)
    driver.get(url)
    accept_privacy_preferences(driver)
    simulate_user_scroll(driver)

    try:
        wait.until(lambda d: d.find_element(By.CSS_SELECTOR, 'span[class*="line-clamp_root"]').text.strip() != "")
        title = driver.find_element(By.CSS_SELECTOR, 'span[class*="line-clamp_root"]').text.strip()
    except Exception:
        title = "Onbekend"

    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class*="product-card-header_unitInfo"]')))
        inhoud_text = driver.find_element(By.CSS_SELECTOR, 'div[class*="product-card-header_unitInfo"]').get_attribute("innerText")
        inhoud = inhoud_text.split("Prijs per")[0].strip()
    except Exception:
        inhoud = "Onbekend"

    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span[aria-hidden="true"]')))
        spans = driver.find_elements(By.CSS_SELECTOR, 'span[aria-hidden="true"]')
        euro = cents = None
        for span in spans:
            if span.text.strip().isdigit():
                if euro is None:
                    euro = span.text.strip()
                else:
                    cents = span.text.strip()
                    break
        prijs = f"€{euro},{cents}" if euro and cents else "Onbekend"
    except Exception:
        prijs = "Onbekend"

    # Debug: save HTML on fail
    if prijs == "Onbekend" or inhoud == "Onbekend":
        with open(f"debug_{index:02d}.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)

    return {
        "Productnaam": title,
        "Inhoud": inhoud,
        "Prijs": prijs,
        "Link": url
    }

product_links = [
    "https://www.ah.nl/producten/product/wi455966/fairtrade-original-community-coffee-aroma-snelfiltermaling", 
"https://www.ah.nl/producten/product/wi455881/fairtrade-original-community-coffee-mild-roast-snelfilter",
"https://www.ah.nl/producten/product/wi457807/fairtrade-original-community-coffee-dark-roast-snelfilter",
"https://www.ah.nl/producten/product/wi455882/fairtrade-original-community-coffee-dark-roast-bonen",
"https://www.ah.nl/producten/product/wi457806/fairtrade-original-colombia-biologische-snelfiltermaling",
"https://www.ah.nl/producten/product/wi457808/fairtrade-original-community-coffee-medium-roast-bonen",
"https://www.ah.nl/producten/product/wi455882/fairtrade-original-community-coffee-dark-roast-bonen",
"https://www.ah.nl/producten/product/wi575568/fairtrade-original-community-coffee-italian-roast-bonen",
"https://www.ah.nl/producten/product/wi455967/fairtrade-original-colombia-biologische-espressobonen",
"https://www.ah.nl/producten/product/wi3738/douwe-egberts-aroma-rood-snelfiltermaling",
"https://www.ah.nl/producten/product/wi3737/douwe-egberts-aroma-rood-snelfiltermaling",
"https://www.ah.nl/producten/product/wi3739/douwe-egberts-aroma-rood-grove-maling-filterkoffie",
"https://www.ah.nl/producten/product/wi33918/douwe-egberts-aroma-rood-grove-maling-filterkoffie",
"https://www.ah.nl/producten/product/wi3735/douwe-egberts-decafe-cafeinevrij-snelfiltermaling",
"https://www.ah.nl/producten/product/wi3663/douwe-egberts-decafe-cafeinevrij-snelfiltermaling",
"https://www.ah.nl/producten/product/wi557751/douwe-egberts-excellent-gold-filtermaling",
"https://www.ah.nl/producten/product/wi1494/perla-huisblends-aroma-snelfiltermaling",
"https://www.ah.nl/producten/product/wi1487/perla-huisblends-aroma-snelfiltermaling",
"https://www.ah.nl/producten/product/wi381295/perla-huisblends-goud-snelfiltermaling",
"https://www.ah.nl/producten/product/wi238687/perla-biologisch-aroma-snelfiltermaling",
"https://www.ah.nl/producten/product/wi3746/kanis-en-gunnink-koffie-snelfiltermaling",
"https://www.ah.nl/producten/product/wi62128/van-nelle-supra-snelfiltermaling-voordeelpak",
"https://www.ah.nl/producten/product/wi457334/lavazza-qualita-rossa-filterkoffie",
"https://www.ah.nl/producten/product/wi457333/lavazza-espresso-italiano-classico-ground-coffee",
"https://www.ah.nl/producten/product/wi457335/lavazza-qualita-oro-ground-coffee",
"https://www.ah.nl/producten/product/wi195679/cafe-intencion-aromatico-snelfiltermaling",
"https://www.ah.nl/producten/product/wi502957/cafe-intencion-fuerte-snelfiltermaling",
"https://www.ah.nl/producten/product/wi164573/douwe-egberts-aroma-rood-bonen",
"https://www.ah.nl/producten/product/wi485609/douwe-egberts-aroma-rood-bonen-voordeelpak",
"https://www.ah.nl/producten/product/wi225657/douwe-egberts-espresso-bonen",
"https://www.ah.nl/producten/product/wi485646/douwe-egberts-espresso-bonen-voordeelpak",
"https://www.ah.nl/producten/product/wi164579/douwe-egberts-excellent-gold-bonen",
"https://www.ah.nl/producten/product/wi438713/douwe-egberts-intens-koffiebonen",
"https://www.ah.nl/producten/product/wi196156/douwe-egberts-excellent-mocca-koffiebonen",
"https://www.ah.nl/producten/product/wi221467/l-or-crema-absolu-classique-coffee-beans",
"https://www.ah.nl/producten/product/wi214571/l-or-espresso-forza-coffee-beans",
"https://www.ah.nl/producten/product/wi214570/l-or-espresso-fortissimo-coffee-beans",
"https://www.ah.nl/producten/product/wi221469/l-or-espresso-onyx-coffee-beans",
"https://www.ah.nl/producten/product/wi161534/perla-huisblends-aroma-koffiebonen",
"https://www.ah.nl/producten/product/wi195720/perla-huisblends-aroma-koffiebonen",
"https://www.ah.nl/producten/product/wi161539/perla-superiore-finest-originale-koffiebonen",
"https://www.ah.nl/producten/product/wi130634/perla-superiore-finest-originale-koffiebonen",
"https://www.ah.nl/producten/product/wi161538/perla-superiore-finest-dark-roast-koffiebonen",
"https://www.ah.nl/producten/product/wi195721/perla-superiore-finest-dark-roast-koffiebonen",
"https://www.ah.nl/producten/product/wi238690/perla-superiore-finest-forte-koffiebonen",
"https://www.ah.nl/producten/product/wi485621/kanis-en-gunnink-medium-roast-koffiebonen-xl-pak",
"https://www.ah.nl/producten/product/wi127407/illy-classico-coffee-beans",
"https://www.ah.nl/producten/product/wi221440/illy-intenso-coffee-beans",
"https://www.ah.nl/producten/product/wi143412/lavazza-qualita-rossa-koffiebonen",
"https://www.ah.nl/producten/product/wi170164/lavazza-espresso-italiano-classico-koffiebonen",
"https://www.ah.nl/producten/product/wi195689/lavazza-qualita-oro-koffiebonen",
"https://www.ah.nl/producten/product/wi170222/cafe-intencion-crema-aromatico-bonen",
"https://www.ah.nl/producten/product/wi395364/cafe-intencion-espresso-intensivo-bonen"
]

klant_namen = [
    "Comm. Coffee Aroma",
"Comm. Coffee Mild snf",
"Comm. Coffee Dark Roast snf",
"Comm. Coffee Decaf snf",
"Comm. Coffee Single Origin snf", 
"Comm. Coffee bonen Medium Roast",
"Comm. Coffee Bonen Dark Roast",
"Comm. Coffee Italian Roast (tot wk19 2024 Aroma Bonen)",
"Comm. Coffee Single Origen Bonen",
"D.E. AROMA  ROOD 250G",
"D.E. AROMA ROOD 500G",
"D.E. AROMA ROOD GROVE MALING 250G",
"D.E. AROMA ROOD GROVE MALING 500G",
"D.E. DECAFE 250G",
"D.E. DECAFE 500G",
"D.E. EXCELLENT 5 250G (GOLD)",
"PERLA AROMA ROOD 250G",
"PERLA AROMA ROOD 500G",
"PERLA GOUD 500G",
"PERLA AROMA BIO 500G",
"KAN.&GUNNINK 500G",
"V.NELLE SUPRA  KOFFIEKRACHT 5 2STx250g",
"LAVAZZA QUALITA ROSSA 250G",
"LAVAZZA ITALIANO CLASSICO ESPRESSO 250G",
"LAVAZZA QUALITA ORO PERFECT  250G",
"CAFE INTENCION ECOLOGICO/ARAMATICO 250G",
"CAFE INTENCION KRACHTIG/FUERTE 250G",
"D.E. AROMA ROOD 500G",
"D.E. AROMA ROOD 1.000G",
"D.E. ESPRESSO 9 500G",
"D.E. ESPRESSO 9 1.000G",
"D.E. EXCELLENT 5 500G",
"D.E. INTENS BONEN 7 500G",
"D.E. MOCCA 7 500G",
"L'OR CREMA CLASSIQUE 500G",
"L'OR FORZA 500G",
"L'OR FORTISSIMO 500G",
"L'OR ONYX 500G",
"PERLA HUISBLEND AROMA BONEN 500G",
"PERLA HUISBLEND AROMA BONEN 1.000G",
"PERLA SUPERIOR ESPRESSO 6 500G",
"PERLA SUPERIOR ESPRESSO 6 1.000G",
"PERLA SUPERIOR DARK ESPRESSO 8 500G",
"PERLA SUPERIOR DARK ESPRESSO 8 1.000G",
"PERLA SUPERIOR ESPRESSO FORTE 9 1.000G",
"KAN.&GUNNINK MEDIUM ROAST 1.000G",
"ILLY ESPRESSO CLASSICO 250G (BLIK)",
"ILLY ESPRESSO DARK ROAST BONEN 250G (BLIK)",
"LAVAZZA QUALITA ROSSA 500G",
"LAVAZZA ITALIANO CLASSICO ESPRESSO 500G",
"LAVAZZA QUALITA ORO PERFECT SYMPH. ESPRESSO 500G",
"CAFE INTENCION KOFFIEBONEN 500G",
"CAFE INTENCION ESPRESSO 500G"
]

filename = f"ah_koffie_{datetime.today().strftime('%Y-%m-%d')}.xlsx"

data = []
batch_size = 10
for i in range(0, len(product_links), batch_size):
    batch = product_links[i:i+batch_size]
    driver = get_driver()
    for j, link in enumerate(batch):
        absolute_index = i + j
        result = scrape_ah_product(driver, link, absolute_index)
        data.append(result)
        print(f"[{absolute_index+1}/{len(product_links)}] ✅ {result['Productnaam']} – {result['Prijs']}")
        time.sleep(random.uniform(30, 60))
    driver.quit()
    if i + batch_size < len(product_links):
        print("🕒 Batch cooling-down (2 min)...")
        time.sleep(120)

for i in range(len(data)):
    if i < len(klant_namen):
        data[i]["Productnaam"] = klant_namen[i]

df = pd.DataFrame(data)
df.to_excel(filename, index=False)

wb = load_workbook(filename)
ws = wb.active
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.hyperlink = cell.value
        cell.font = Font(color="0000EE", underline="single")
wb.save(filename)

try:
    verzend_excel_via_mail(
        bestandspad=filename,
        ontvangers=EMAIL_RECEIVERS,
        afzender=EMAIL_USER,
        wachtwoord=EMAIL_PASS
    )
except Exception as e:
    print(f"[ERROR] E-mail verzenden mislukt: {e}")
    sys.exit(2)
else:
    print("[INFO] De scraper is klaar en de koffie-resultaten zijn gemaild.")
    sys.exit(0)
