# ========================================
# 🎟️ FTO WINKELSCRAPER – DIRK (ENV, GUI, per product driver, automation)
# Versie: 1.3c – Fred Luijkx, Netflex BV
# ========================================

import sys
import time
import random
import pandas as pd
from datetime import datetime
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font
import smtplib
from email.message import EmailMessage
import os
from dotenv import load_dotenv

# =============== PAS DIT PAD AAN AAN JOUW VM! =================
chrome_path = r"C:\chrome-testing\chrome-win64\chrome.exe"
chromedriver_path = r"C:\chrome-testing\chromedriver-win64\chromedriver.exe"
# ==============================================================

# .env inladen
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

dotenv_path = resource_path(".env")
load_dotenv(dotenv_path="C:/scrapers/.env")

EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASSWORD")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")  # Mag komma-gescheiden zijn

if not EMAIL_USER or not EMAIL_PASS or not EMAIL_RECEIVER:
    print("[FATAL] Vul het bestand .env aan met: EMAIL_USER, EMAIL_PASSWORD, EMAIL_RECEIVER. Script stopt nu.")
    sys.exit(1)
EMAIL_RECEIVERS = [r.strip() for r in EMAIL_RECEIVER.split(",")]

# 📧 Excel mailen
def verzend_excel_via_mail(bestandspad, ontvangers, afzender, wachtwoord):
    msg = EmailMessage()
    msg["Subject"] = "🎟️ FTO – Dirk scraperresultaten"
    msg["From"] = afzender
    msg["To"] = ", ".join(ontvangers)
    msg.set_content("Zie bijlage voor de gescrapete producten van de Dirk.")

    with open(bestandspad, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(bestandspad)
        )

    with smtplib.SMTP("smtp.office365.com", 587) as smtp:
        smtp.starttls()
        smtp.login(afzender, wachtwoord)
        smtp.send_message(msg)

# 🔗 Dirk-productlinks
product_links = [
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/fairtrade%20filterkoffie%20aroma/74329",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/fairtrade%20filterkoffie%20mild/74330",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/fairtrade%20filterkoffie%20decafe%20cafeinevrij/74331",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20aroma%20rood%20filterkoffie/3460",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20aroma%20rood%20filterkoffie/3826",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20aroma%20rood%20%20filterkoffie%20grove%20maling/2694",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20d%c3%a9caf%c3%a9%20filterkoffie/9119",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20filterkoffie%20excellent%20gold%20sterkte%205/22403",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/1%20de%20beste%20filterkoffie%20roodmerk/84208",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/1%20de%20beste%20filterkoffie%20roodmerk/84209",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/kanis%20%26%20gunnink%20koffie%20snelfiltermaling/9354",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/van%20nelle%20filterkoffie/20784",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lavazza%20snelfiltermaling%20qualita%20rossa/6176",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lavazza%20snelfilterkoffie%20caf%c3%a9%20espresso/1720",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lavazza%20qualita%20oro/31792",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/bio%2b%20filterkoffie%20dutch%20roast/56379",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20aroma%20rood%20koffiebonen/1099",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20koffiebonen%20aroma%20rood/65563",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20koffiebonen%20espresso/65835",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20aroma%20variaties%20excellent%20koffiebonen/1225",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20intens%20koffiebonen/39791",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/douwe%20egberts%20aroma%20variaties%20mocca%20koffiebonen/10475",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lor%20espresso%20forza%20koffiebonen/6615",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lor%20espresso%20fortissimo%20koffiebonen/5585",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lor%20espresso%20onyx%20koffiebonen/7846",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/1%20de%20beste%20koffiebonen%20rood/84670",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/1%20de%20beste%20koffiebonen%20dark%20espresso/84671",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/kanis%20%26%20gunnink%20koffiebonen%20medium%20roast/65503",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/illy%20bonen%20classico/82394",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lavazza%20espressobonen/1557",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/lavazza%20espressobonen%20qualit%c3%a0%20oro/3589",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/bio%2b%20koffiebonen%20krachtig/57127",
    "https://www.dirk.nl/boodschappen/voorraadkast/internationale-keuken/fairtrade%20kokosmelk/24548",
    "https://www.dirk.nl/boodschappen/voorraadkast/internationale-keuken/fairtrade%20kruidenpasta%20gele%20curry/42039",
    "https://www.dirk.nl/boodschappen/voorraadkast/internationale-keuken/inproba%20kruidenpasta%20groene%20curry/65459",
    "https://www.dirk.nl/boodschappen/voorraadkast/internationale-keuken/inproba%20kruidenpasta%20rode%20curry/66126",
    "https://www.dirk.nl/boodschappen/voorraadkast/internationale-keuken/inproba%20kruidenpasta%20gele%20curry/79832",
    "https://www.dirk.nl/boodschappen/voorraadkast/internationale-keuken/koh%20thai%20curry%20pasta%20groen/5117",
    "https://www.dirk.nl/boodschappen/brood-beleg-koek/broodbeleg/de%20ruijter%20hagelslag%20puur/83890",
    "https://www.dirk.nl/boodschappen/brood-beleg-koek/broodbeleg/venz%20hagelslag%20puur/83893",
    "https://www.dirk.nl/boodschappen/brood-beleg-koek/broodbeleg/1%20de%20beste%20hagelslag%20puur/75810",
    "https://www.dirk.nl/boodschappen/voorraadkast/speciale-voeding/c%c3%a9r%c3%a9al%20hagelslag%20puur/96402",
    "https://www.dirk.nl/boodschappen/dranken-sap-koffie-thee/koffie-cacao/blooker%20cacaopoeder/102611",
    "https://www.dirk.nl/boodschappen/voorraadkast/speciale-voeding/raw%20organic%20food%20cacaopoeder/63635",
    "https://www.dirk.nl/boodschappen/aardappelen-groente-fruit/fruitconserven/del%20monte%20ananasschijven%20op%20sap/9109",
    "https://www.dirk.nl/boodschappen/aardappelen-groente-fruit/fruitconserven/del%20monte%20ananasblokjes%20op%20sap/9559",
    "https://www.dirk.nl/boodschappen/aardappelen-groente-fruit/fruitconserven/1%20de%20beste%20ananasschijven%20op%20sap/59697",
    "https://www.dirk.nl/boodschappen/aardappelen-groente-fruit/fruitconserven/1%20de%20beste%20ananasstukjes%20op%20sap/62194",
]

klant_namen = [
    "LIRP Koffie Aroma snf, MH, 250g",
    "LIRP Koffie Mild snf, bio, MH, 250g",
    "LIRP Koffie Décaf snf, MH, 250g", 
    "D.E. AROMA  ROOD 250G",
    "D.E. AROMA ROOD 500G",
    "D.E. AROMA ROOD GROVE MALING 250G",
    "D.E. DECAFE 250G",
    "D.E. EXCELLENT 5 250G",
    "1 DE BESTE AROMA ROOD 250G ",
    "1 DE BESTE AROMA ROOD 500G ",
    "KAN.&GUNNINK 500G",
    "V.NELLE SUPRA  KOFFIEKRACHT 5 2STx250g",
    "LAVAZZA QUALITA ROSSA 250G",
    "LAVAZZA ITALIANO CLASSICO ESPRESSO 250G",
    "LAVAZZA QUALITA ORO PERFECT SYMPH. ESPRESSO 250G",
    "BIO+ AROMA/DUTCH ROAST 250G",
    "D.E. AROMA ROOD 500G",
    "D.E. AROMA ROOD 1.000G",
    "D.E. ESPRESSO 9 1.000G",
    "D.E. EXCELLENT 5 500G",
    "D.E. INTENS BONEN 7 500G",
    "D.E. MOCCA 7 500G",
    "L'OR FORZA 500G",
    "L'OR FORTISSIMO 500G",
    "L'OR ONYX 500G",
    "1 DE BESTE AROMA BONEN 5 1.000G",
    "1 DE BESTE DARK ROAST ESPRESSO 1.000G", 
    "KAN.&GUNNINK MEDIUM ROAST 1.000G",
    "ILLY ESPRESSO CLASSICO 250G (BLIK)",
    "LAVAZZA CAFFE ESPRESSO 500G",
    "LAVAZZA QUALITA ORO PERFECT SYMPH. ESPRESSO 500G",
    "BIO+ ESPRESSO 450G / KOFFIEBONEN KRACHTIG",
    "FAIRTRADE ORIGINAL 400ML",
    "FAIRTRADE ORIGINAL GELE CURRY",
    "INPROBA GROENE CURRY",
    "INPROBA RODE CURRY",
    "INPROBA GELE CURRY",
    "KOH THAI POT",
    "DE RUIJTER PUUR, 390G",
    "VENZ, 400G",
    "1 DE BESTE, 600G",
    "CEREAL, 200G",
    "BLOOKER CACAOPOEDER (ook fairtrade sinds ergens in 2024)",
    "RAW CACAOPOEDER, BIOLOGISCH, 100G",
    "DEL MONTE ANANASSCHIJVEN OP SAP, 220G",
    "DEL MONTE ANANASSTUKJES OP SAP, 230G",
    "1 DE BESTE ANANASSCHIJVEN OP SAP, 567G",
    "1 DE BESTE ANANASSTUKJES OP SAP, 227G",
]

MAX_RETRIES = 3

def scrape_dirk_product(url):
    for attempt in range(1, MAX_RETRIES + 1):
        options = uc.ChromeOptions()
        options.binary_location = chrome_path
        # HEADLESS UITGECOMMENTEERD (voor debug/anti-bot): 
        # options.add_argument('--headless=new')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-software-rasterizer')
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        try:
            driver = uc.Chrome(options=options, driver_executable_path=chromedriver_path)
            wait = WebDriverWait(driver, 40)
            time.sleep(random.uniform(2, 4))  # Klein beetje wachten tegen anti-bot
            driver.get(url)
            title = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.title > h1'))).text.strip()
            inhoud = driver.find_element(By.CSS_SELECTOR, 'p.subtitle').text.strip()
            prijs = "Onbekend"
            try:
                euro = driver.find_element(By.CLASS_NAME, "price-large").text.strip()
                cent_elements = driver.find_elements(By.CLASS_NAME, "price-small")
                if cent_elements:
                    prijs = f"€{euro},{cent_elements[0].text.strip()}"
                elif euro.isdigit():
                    prijs = f"€0,{euro}"
            except Exception:
                pass
            driver.quit()
            return {
                "Productnaam": title,
                "Inhoud": inhoud,
                "Prijs": prijs,
                "Link": url
            }
        except Exception as e:
            print(f"[Retry {attempt}/{MAX_RETRIES}] Fout bij ophalen: {url}\n{e}")
            try:
                driver.quit()
            except:
                pass
            time.sleep(5 * attempt)
    print(f"[❌] Product niet gescraped: {url}")
    return {"Productnaam": "Mislukt", "Inhoud": "", "Prijs": "", "Link": url}

# 📅 Excel-bestandsnaam
today = datetime.today().strftime("%Y-%m-%d")
filename = f"dirk_producten_{today}.xlsx"

results = []
for idx, link in enumerate(product_links, 1):
    print(f"🛒 ({idx}/{len(product_links)}) Ophalen: {link}")
    results.append(scrape_dirk_product(link))
    time.sleep(random.uniform(3, 5))

for i in range(min(len(results), len(klant_namen))):
    results[i]["Productnaam"] = klant_namen[i]

pd.DataFrame(results).to_excel(filename, index=False)

wb = load_workbook(filename)
ws = wb.active
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.hyperlink = cell.value
        cell.font = Font(color="0000EE", underline="single")
wb.save(filename)

# 📧 Mail versturen via .env
try:
    verzend_excel_via_mail(
        bestandspad=filename,
        ontvangers=EMAIL_RECEIVERS,
        afzender=EMAIL_USER,
        wachtwoord=EMAIL_PASS
    )
except Exception as e:
    print(f"[ERROR] E-mail verzenden mislukt: {e}")
    sys.exit(2)

print(f"✅ Scraping en e-mail voltooid: {filename}")
sys.exit(0)


